from abc import ABC, abstractmethod
from typing import List
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string
import csv
import os

class FileReader(ABC):
    @abstractmethod
    def extract_text(self, file_path, **kwargs) -> List[str]:
        pass

class FileWriter(ABC):
    @abstractmethod
    def write_text(self, file_path, texts, **kwargs) -> bool:
        pass

class ExcelFileReader(FileReader):
    def extract_text(self, file_path, target_column, start_row, end_row) -> List[str]:
        start_row = int(start_row)
        end_row = int(end_row)
        
        workbook = load_workbook(file_path)
        sheet = workbook.active
        texts = []

        # Convert the target_column letter to an index
        target_column_index = column_index_from_string(target_column)

        for row in range(start_row, end_row + 1):
            cell_value = sheet.cell(row=row, column=target_column_index).value
            if cell_value is not None:
                texts.append(str(cell_value))
            else:
                texts.append("")

        return texts

    
class CSVFileReader(FileReader):
    def extract_text(self, file_path, target_column, start_row, end_row) -> List[str]:
        texts = []
        with open(file_path, 'r', newline='') as csvfile:
            reader = csv.reader(csvfile)
            for row_number, row in enumerate(reader, start=1):
                if start_row <= row_number <= end_row and target_column < len(row):
                    cell_value = row[target_column]
                    if cell_value:
                        texts.append(str(cell_value))
        return texts
    
class ExcelFileWriter(FileWriter):
    def write_text(self, file_path, texts, start_column, start_row, end_row) -> bool:
        assert end_row - start_row + 1 == len(texts)
        start_row = int(start_row)
        end_row = int(end_row)
        start_column = column_index_from_string(start_column)
        try:
            try:
                workbook = load_workbook(file_path)
            except FileNotFoundError:
                workbook = Workbook()
            # Create a copy of the active sheet
            original_sheet = workbook.active
            copied_sheet = workbook.copy_worksheet(original_sheet)
            copied_sheet.title = "translated"
            sheet = workbook.active
            language_type = len(texts[0])
            for i in range(start_row, end_row+1):
                for j in range(language_type):
                    sheet.cell(row=i, column=start_column + j, value=texts[i-start_row][j]['generated_translation'])
            directory, original_filename = os.path.split(file_path)
            new_filename = original_filename.replace(".", f"_{start_row}_{end_row}_translated.")
            new_file_path = os.path.join(directory, new_filename)
            print("new_file_path", new_file_path)
            workbook.save(new_file_path)
            return new_file_path

        except Exception as e:
            raise FileExistsError(f"Error writing to Excel: {e}")
        
class FileReaderFactory:
    @staticmethod
    def create_reader(file_path):
        _, extension = os.path.splitext(file_path)from abc import ABC, abstractmethod
from typing import List
from openpyxl import Workbook, load_workbook
import xlrd
from openpyxl.utils import column_index_from_string
import csv
import os

class FileReader(ABC):
    @abstractmethod
    def extract_text(self, file_path, **kwargs) -> List[str]:
        pass

class FileWriter(ABC):
    @abstractmethod
    def write_text(self, file_path, texts, **kwargs) -> bool:
        pass

class ExcelFileReader(FileReader):
    def extract_text(self, file_path, target_column, start_row, end_row) -> List[str]:
        start_row = int(start_row)
        end_row = int(end_row)
        
        workbook = load_workbook(file_path)
        sheet = workbook.active
        texts = []

        # Convert the target_column letter to an index
        target_column_index = column_index_from_string(target_column)

        for row in range(start_row, end_row + 1):
            cell_value = sheet.cell(row=row, column=target_column_index).value
            if cell_value is not None:
                texts.append(str(cell_value))
            else:
                texts.append("")

        return texts

    
class CSVFileReader(FileReader):
    def extract_text(self, file_path, target_column, start_row, end_row) -> List[str]:
        texts = []
        with open(file_path, 'r', newline='') as csvfile:
            reader = csv.reader(csvfile)
            for row_number, row in enumerate(reader, start=1):
                if start_row <= row_number <= end_row and target_column < len(row):
                    cell_value = row[target_column]
                    if cell_value:
                        texts.append(str(cell_value))
        return texts
    
class ExcelFileWriter(FileWriter):
    def write_text(self, file_path, texts, start_column, start_row, end_row) -> bool:
        assert end_row - start_row + 1 == len(texts)
        start_row = int(start_row)
        end_row = int(end_row)
        start_column = column_index_from_string(start_column)
        try:
            try:
                workbook = load_workbook(file_path)
            except FileNotFoundError:
                workbook = Workbook()
            # Create a copy of the active sheet
            original_sheet = workbook.active
            copied_sheet = workbook.copy_worksheet(original_sheet)
            copied_sheet.title = "translated"
            sheet = workbook.active
            language_type = len(texts[0])
            for i in range(start_row, end_row+1):
                for j in range(language_type):
                    sheet.cell(row=i, column=start_column + j, value=texts[i-start_row][j]['generated_translation'])
            directory, original_filename = os.path.split(file_path)
            new_filename = original_filename.replace(".", f"_{start_row}_{end_row}_translated.")
            new_file_path = os.path.join(directory, new_filename)
            print("new_file_path", new_file_path)
            workbook.save(new_file_path)
            return new_file_path

        except Exception as e:
            raise FileExistsError(f"Error writing to Excel: {e}")
        

class FileReaderFactory:
    @staticmethod
    def create_reader(file_path):
        _, extension = os.path.splitext(file_path)
        extension = extension.lower()

        if extension == '.xls':
            # Convert .xls to .xlsx
            file_path = convert_xls_to_xlsx(file_path)
            extension = '.xlsx'

        reader_mapping = {
            '.xlsx': ExcelFileReader,
            '.csv': CSVFileReader,
        }

        reader_class = reader_mapping.get(extension)

        if reader_class:
            return reader_class(), file_path
        else:
            raise ValueError(f"Unsupported file format: {extension}")
        
    @staticmethod
    def count_rows(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
        count = len([row for row in sheet if not all([cell.value is None for cell in row])])
        return count

def convert_xls_to_xlsx(file_path):
    workbook_xls = xlrd.open_workbook(file_path)
    sheet_xls = workbook_xls.sheet_by_index(0)

    workbook_xlsx = Workbook()
    sheet_xlsx = workbook_xlsx.active

    # Copy data from .xls to .xlsx
    for row in range(sheet_xls.nrows):
        for col in range(sheet_xls.ncols):
            cell_value = sheet_xls.cell_value(row, col)
            sheet_xlsx.cell(row=row+1, column=col+1, value=cell_value)

    # Save the new .xlsx file
    new_file_path = file_path.replace('.xls', '.xlsx')
    workbook_xlsx.save(new_file_path)
    return new_file_path
        extension = extension.lower()

        reader_mapping = {
            '.xlsx': ExcelFileReader,
            '.csv': CSVFileReader,
        }

        reader_class = reader_mapping.get(extension)

        if reader_class:
            return reader_class()
        else:
            raise ValueError(f"Unsupported file format: {extension}")
