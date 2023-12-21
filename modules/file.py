from abc import ABC, abstractmethod
from typing import List
from openpyxl import Workbook, load_workbook
import csv
import os

class FileReader(ABC):
    @abstractmethod
    def extract_text(self, file_path, **kwargs) -> List[str]:
        pass

class FileWriter(ABC):
    @abstractmethod
    def write_text(self, file_path, texts, **kwargs) -> bool:
        pass

class ExcelFileReader(FileReader):
    def extract_text(self, file_path, target_column, start_row, end_row) -> List[str]:
        workbook = load_workbook(file_path)
        sheet = workbook.active  # Assuming we are working with the active sheet
        texts = []
        for row in range(start_row, end_row + 1):
            cell_value = sheet.cell(row=row, column=target_column).value
            if cell_value is not None:
                texts.append(str(cell_value))
        return texts
    
class CSVFileReader(FileReader):
    def extract_text(self, file_path, target_column, start_row, end_row) -> List[str]:
        texts = []
        with open(file_path, 'r', newline='') as csvfile:
            reader = csv.reader(csvfile)
            for row_number, row in enumerate(reader, start=1):
                if start_row <= row_number <= end_row and target_column < len(row):
                    cell_value = row[target_column]
                    if cell_value:
                        texts.append(str(cell_value))
        return texts
        
    
class ExcelFileWriter(FileWriter):
    def write_text(self, file_path, texts, start_row, end_row, target_column) -> bool:
        try:
            try:
                workbook = load_workbook(file_path)
            except FileNotFoundError:
                workbook = Workbook()
            sheet = workbook.active
            for row_num, text in enumerate(texts, start=start_row):
                sheet.cell(row=row_num, column=target_column, value=text)
            workbook.save(file_path)
            return True

        except Exception as e:
            print(f"Error writing to Excel: {e}")
            return False
        
class FileReaderFactory:
    @staticmethod
    def create_reader(file_path):
        _, extension = os.path.splitext(file_path)
        extension = extension.lower()

        reader_mapping = {
            '.xlsx': ExcelFileReader,
            '.csv': CSVFileReader,
        }

        reader_class = reader_mapping.get(extension)

        if reader_class:
            return reader_class()
        else:
            raise ValueError(f"Unsupported file format: {extension}")